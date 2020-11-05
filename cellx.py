from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from datetime import datetime

import os

d = datetime.now()

headFontStyle = Font(name="Arial", size = "30", bold=True)
addressFontStyle = Font(name="Arial", size = "14", bold=True)
rowFontStyle = Font(name="Arial", size = "12", bold=False)

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

center = Alignment(horizontal="center", vertical="center")

print("Loading file...")
wb = load_workbook("cell.xlsx")
ws = wb.worksheets[0]
print("Done")

'''
A - артикул (3 merged cells)
B - наименование товара
C - адресом
D - колличество(шт)
ws.max_row - количество строк на листе
'''

raw_data = {}

print("Collecting data...")

rowLen = max((c.row for c in ws['A'] if c.value is not None))

for row in ws.iter_rows(min_row=1, min_col=1, max_row=rowLen, max_col=4):
		art = row[0].value
		name = row[1].value
		adr = row[2].value
		q = row[3].value
		if q == None:
			q = 0
		if adr in raw_data:
			raw_data[adr].append([name, art, q])
		else:
			raw_data[adr] = [[name, art, q]]
			
print("raw_data collected.")
wb.close()

#DAC45850041A48-KIT  14-12-02  18

adr_list = list(raw_data)
adr_list.sort()

adressless_l = [] # Wrong adress
for key in raw_data:
	if key[2:4].isdigit() == False:
		adressless_l.append(key)
		
adressless_d = {}
for i in adressless_l:
	af = raw_data.pop(i)
	adressless_d.update({i: af})
	
adr_list = list(raw_data)
adr_list.sort()

os.mkdir(f"{d}")
os.mkdir(f"{d}/Без адреса")

for i in range(1, 21):
	os.mkdir(f"{d}/{i} ряд")

def creds(adress):
	"""ф-я рассчёта папки для сохранения, оглавления для стелажки/название файла используя указанный адрес"""
	file_name = f"{adress[2: 4]} - {adress[5: 7]}"
	folder = f"{int(adress[2: 4])} ряд"
	return [file_name, folder]

def stak_create(dictionary):
	print("Creating staks...")
	"""Функция, формирующая стелажку из указанного словаря"""
	r = 1
	for a in dictionary:
		af = raw_data[a]
		row_number = creds(a)[1]
		stak_title = creds(a)[0]
		if os.path.isfile(f"{d}/{row_number}/{stak_title}.xlsx") == False:
			r = 1
			wb = Workbook()
			ws = wb.active
			ws.title = stak_title
			ws.merge_cells('A1:C1')
			ws.cell(row=1, column=1, value=stak_title).font = headFontStyle
			ws["A1"].alignment = center
			ws.column_dimensions['A'].width = 50
			ws.column_dimensions['B'].width = 19
			ws.column_dimensions['C'].width = 6
			ws.row_dimensions[1].height = 40
			r += 1
			wb.save(f"{d}/{row_number}/{stak_title}.xlsx")
		
		wb = load_workbook(f"{d}/{row_number}/{stak_title}.xlsx")
		ws = wb.worksheets[0]
		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
		ws.cell(row=r, column=1, value=a).border = thick_border
		ws.cell(row=r, column=1).alignment = center
		ws.cell(row=r, column=1).font = addressFontStyle
		r += 1
		for row in af:
			ws.append(row)
			for n in range(1, 4):
		 		ws.cell(row=r, column=n).border = thin_border
		 		ws.cell(row=r, column=n).font = rowFontStyle
			r += 1
		wb.save(f"{d}/{row_number}/{stak_title}.xlsx")
	print("Staks created")

stak_create(adr_list)

def wrong_adress():
	print("Searching wrong adresses...")
	wb = Workbook()
	ws = wb.active
	ws.title = "Без адреса"
	r = 1
	c = 1
	ws.merge_cells('A1:C1')
	ws.cell(row=1, column=1, value="Без адреса").font = headFontStyle
	ws["A1"].alignment = center
	ws.column_dimensions['A'].width = 50
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 6
	ws.row_dimensions[1].height = 40
	r += 1
	for i in adressless_l:
		 af = adressless_d[i]
		 ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
		 ws.cell(row=r, column=c).value = i
		 ws.cell(row=r, column=c).border = thick_border
		 ws.cell(row=r, column=c).alignment = center
		 r += 1
		 for row in af:
		 	ws.append(row)
		 	for n in range(1, 4):
		 		ws.cell(row=r, column=n).border = thin_border
		 	r += 1
	wb.save(f"{d}/Без адреса/Без адреса.xlsx")
	print("Wrong sdresses found")

wrong_adress()


print("Program finished")